#sales_router/src/sales_routing/reporting/export_cluster_summary.py

# ============================================================
# 📊 src/sales_routing/reporting/export_cluster_summary.py
# ============================================================

import io
import os
from pathlib import Path
import pandas as pd
from loguru import logger
from src.database.db_connection import get_connection
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter

BASE_OUTPUT = Path("/app/output/reports")  # caminho absoluto no container


def routing_resumo_to_bytes(
    tenant_id: int, routing_id: str, schema: str = "public"
) -> bytes:
    """Gera o XLSX em memória e retorna os bytes (sem persistir em disco).
    Usado pelo endpoint /relatorio/resumo via StreamingResponse.
    `schema`: 'operacional' resolve as views no schema operacional."""
    buffer = io.BytesIO()
    if not _escrever_resumo(tenant_id, routing_id, buffer, schema=schema):
        raise ValueError("Nenhum dado encontrado para a roteirização informada.")
    buffer.seek(0)
    return buffer.getvalue()


def exportar_resumo_cluster(tenant_id: int, routing_id: str):
    """CLI standalone: grava em disco (uso manual via python -m).
    A API agora usa routing_resumo_to_bytes."""
    logger.info(
        f"📊 Exportando resumo XLSX | tenant={tenant_id} | routing_id={routing_id}"
    )
    pasta = BASE_OUTPUT / str(tenant_id)
    pasta.mkdir(parents=True, exist_ok=True)
    arquivo = pasta / f"routing_resumo_{routing_id}.xlsx"
    with open(arquivo, "wb") as f:
        if not _escrever_resumo(tenant_id, routing_id, f):
            return None
    logger.success(f"✅ XLSX salvo em: {arquivo}")
    return str(arquivo)


def _escrever_resumo(
    tenant_id: int, routing_id: str, output, schema: str = "public"
) -> bool:
    """Núcleo: lê do DB e escreve XLSX em qualquer file-like.
    Retorna False se não houver dados."""

    sql = """
        SELECT
            cluster_id              AS "Cluster",
            qtd_pdvs                AS "PDVs",
            tempo_medio_min         AS "Tempo médio (min)",
            tempo_max_min           AS "Tempo máximo (min)",
            tempo_total_min         AS "Tempo total (min)",
            dist_media_km           AS "Distância média (km)",
            dist_max_km             AS "Distância máxima (km)",
            dist_total_km           AS "Distância total (km)",
            valor_total_vendas      AS "Valor total vendas",
            centro_lat              AS "Latitude centro",
            centro_lon              AS "Longitude centro"
        FROM vw_sales_routing_resumo_cluster
        WHERE tenant_id = %s
          AND routing_id = %s
        ORDER BY cluster_id;
    """

    conn = get_connection()
    try:
        if schema and schema != "public":
            with conn.cursor() as _c:
                _c.execute(f"SET search_path TO {schema}, public")
        df = pd.read_sql_query(sql, conn, params=(tenant_id, routing_id))

        if df.empty:
            logger.warning("⚠️ Nenhum dado encontrado para exportação.")
            return False

        with pd.ExcelWriter(output, engine="openpyxl") as writer:
            df.to_excel(writer, sheet_name="Resumo por Cluster", index=False)
            ws = writer.book["Resumo por Cluster"]

            ws.freeze_panes = "A2"
            header_font = Font(bold=True)
            header_align = Alignment(horizontal="center", vertical="center")
            for cell in ws[1]:
                cell.font = header_font
                cell.alignment = header_align

            widths = {1: 10, 2: 8, 3: 20, 4: 22, 5: 22, 6: 22, 7: 24, 8: 24, 9: 22, 10: 18, 11: 18}
            for col_idx, width in widths.items():
                ws.column_dimensions[get_column_letter(col_idx)].width = width

        return True

    finally:
        conn.close()



if __name__ == "__main__":
    import argparse

    parser = argparse.ArgumentParser(description="Exporta resumo de clusters (Sales Routing).")
    parser.add_argument("--tenant", type=int, required=True, help="Tenant ID")
    parser.add_argument("--routing_id", type=str, required=True, help="UUID da roteirização")
    args = parser.parse_args()

    caminho = exportar_resumo_cluster(args.tenant, args.routing_id)
    if caminho:
        print(f"\n📂 Arquivo gerado: {caminho}\n")
