#sales_router/src/sales_clusterization/reporting/export_cluster_resumo_xlsx.py

# ============================================================
# 📦 src/sales_clusterization/reporting/export_cluster_resumo_xlsx.py
# ============================================================

import io
import os
import pandas as pd
import argparse
from loguru import logger
from database.db_connection import get_connection


def cluster_resumo_to_bytes(tenant_id: int, clusterization_id: str) -> bytes:
    """Gera o XLSX em memória e retorna os bytes (sem persistir em disco).
    Usado pelo endpoint de download via StreamingResponse."""
    buffer = io.BytesIO()
    _escrever_resumo(tenant_id, clusterization_id, buffer)
    buffer.seek(0)
    return buffer.getvalue()


def exportar_cluster_resumo(tenant_id: int, clusterization_id: str):
    """CLI standalone: grava em disco (uso manual via python -m)."""
    logger.info(
        f"📊 Exportando resumo de clusters | tenant={tenant_id} | clusterization_id={clusterization_id}"
    )

    output_dir = f"output/reports/{tenant_id}"
    os.makedirs(output_dir, exist_ok=True)
    output_path = os.path.join(output_dir, f"cluster_resumo_{clusterization_id}.xlsx")

    with open(output_path, "wb") as f:
        _escrever_resumo(tenant_id, clusterization_id, f)
    logger.success(f"✅ Excel salvo em: {output_path}")


def _escrever_resumo(tenant_id: int, clusterization_id: str, output):
    """Núcleo: lê do DB e escreve XLSX em qualquer file-like (BytesIO ou file)."""
    conn = get_connection()

    # 🔍 Run mais recente
    query_run = f"""
        SELECT id AS run_id
        FROM cluster_run
        WHERE tenant_id = {tenant_id}
          AND clusterization_id = '{clusterization_id}'
        ORDER BY criado_em DESC
        LIMIT 1;
    """
    run_df = pd.read_sql_query(query_run, conn)

    if run_df.empty:
        conn.close()
        raise ValueError("❌ Nenhum run encontrado")

    run_id = int(run_df.iloc[0]["run_id"])

    # 📊 Dados
    query = f"""
        SELECT *
        FROM v_cluster_resumo
        WHERE tenant_id = {tenant_id}
          AND run_id = {run_id}
        ORDER BY cluster_label;
    """
    df = pd.read_sql_query(query, conn)
    conn.close()

    if df.empty:
        raise ValueError("❌ Nenhum dado encontrado")

    # 🚫 Remove coluna técnica
    df = df.drop(columns=["metrics_json"], errors="ignore")

    # 🏷️ Renomeio executivo
    df = df.rename(
        columns={
            "cluster_label": "Cluster",
            "n_pdvs": "PDVs",
            "dist_media_km": "Distância média (km)",
            "dist_max_km": "Distância máxima (km)",
            "tempo_medio_min": "Tempo médio (min)",
            "tempo_max_min": "Tempo máximo (min)",
            "centro_lat": "Latitude centro",
            "centro_lon": "Longitude centro",
        }
    )

    # 📐 Ordem final
    df = df[
        [
            "Cluster",
            "PDVs",
            "Distância média (km)",
            "Distância máxima (km)",
            "Tempo médio (min)",
            "Tempo máximo (min)",
            "Latitude centro",
            "Longitude centro",
        ]
    ]

    from openpyxl.styles import Font, Alignment
    from openpyxl.utils import get_column_letter

    # ===============================
    # 🧾 Escrita EXECUTIVA no Excel
    # ===============================
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="Resumo por Cluster", index=False)

        ws = writer.book["Resumo por Cluster"]

        # 🔒 Freeze header
        ws.freeze_panes = "A2"

        # 🎨 Header (negrito + centralizado)
        header_font = Font(bold=True)
        header_align = Alignment(horizontal="center", vertical="center")

        for cell in ws[1]:
            cell.font = header_font
            cell.alignment = header_align

        # 📐 Largura das colunas
        widths = {
            1: 10,   # Cluster
            2: 8,    # PDVs
            3: 22,   # Distância média
            4: 24,   # Distância máxima
            5: 20,   # Tempo médio
            6: 20,   # Tempo máximo
            7: 18,   # Latitude
            8: 18,   # Longitude
        }

        for col_idx, width in widths.items():
            ws.column_dimensions[get_column_letter(col_idx)].width = width


if __name__ == "__main__":
    parser = argparse.ArgumentParser()
    parser.add_argument("--tenant_id", type=int, required=True)
    parser.add_argument("--clusterization_id", type=str, required=True)
    args = parser.parse_args()

    exportar_cluster_resumo(args.tenant_id, args.clusterization_id)
